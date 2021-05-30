import copy
import cvxpy as cp
import numpy as np
import math
import networkx as nx
import numpy as np
from graphviz import Digraph
from openpyxl import load_workbook
import openpyxl
from networkx.readwrite import json_graph
from networkx.drawing.nx_agraph import write_dot, graphviz_layout
import matplotlib.pyplot as plt
import pylab
from openpyxl.styles import Color, Fill,Font,PatternFill,Alignment

# кароч 1 связи, которые проставил пользователь ниже функция пример
# выходной список - список индексов, по которым нужно будет строить граф - учебник фрола - сошлось - ок
def determite_work(connection_from_user):
    rank_table = []
    print(connection_from_user)
    data_info = list(range(len(connection_from_user)))
    data_info.insert(0,-1)
    info_about_rank = {-1: set()}
    not_none_index = []
    for i in range(len(connection_from_user)):
        if len(connection_from_user[i]) == 1 and connection_from_user[i][0] == -1:
            info_about_rank.get(-1).add(i)
        else:
            not_none_index.append(i)
    print([connection_from_user[i] for i in not_none_index])

    # print(not_none_index)
    set_insected = set()
    prom = []
    print(not_none_index)
    while len(not_none_index) != 0:
        for i in not_none_index:
            last_rank = -20
            check_index_valide = 0
            for j in connection_from_user[i]:
                for i1, i2 in info_about_rank.items():
                    if j in i2:
                        check_index_valide += 1
                        if last_rank < i1:
                            last_rank = i1

            if check_index_valide == len(connection_from_user[i]):
                # print(connection_from_user[i])
                # print(i)
                if not last_rank + 1 in info_about_rank.keys():
                    info_about_rank[last_rank + 1] = set()
                    info_about_rank.get(last_rank + 1).add(i)
                else:
                    info_about_rank.get(last_rank + 1).add(i)
            else:
                prom.append(i)
            # print('----')

            # print(info_about_rank)
                # print(last_rank)
                # print(j)
                # print()
        # print(info_about_rank)

        not_none_index = copy.deepcopy(prom)
        prom = []
    for i1, i in info_about_rank.items():
        print(i1, '->', i)
    return info_about_rank

def find_t_and_T(connection_from_user, ranked_info, info_about_work):
    prom_info_date = []
    dict_info_t = {}
    k = 0
    for i in ranked_info.values():
        prom_info = []
        for j in i:
            if connection_from_user[j][0] == -1:
                dict_info_t[f't_{j}'] = info_about_work[j]
                dict_info_t[f'T_{j}'] = info_about_work[j]
                k += 1
            else:
                # print('---------')
                # print(i)
                # print(j)
                # print(dict_info_t)
                # print(connection_from_user[j])
                dict_info_t[f't_{j}'] = info_about_work[j]
                prom_list = [dict_info_t.get(f'T_{i}') for i in connection_from_user[j]]
                # print('---------')

                # # for i in connection_from_user:
                # if f'T_{k}' == 'T_11':
                #     print('------')
                #     print(dict_info_mini_t)
                #     print(connection_from_user[j])
                #     print(prom_list)
                #     print('------')
                dict_info_t[f'T_{j}'] = info_about_work[j] + max(prom_list)
                k += 1
        prom_info_date.append(copy.deepcopy(prom_info))
        prom_info.clear()
    # print(dict_info_big_T)
    # print(dict_info_mini_t)
    print(dict_info_t)
    return dict_info_t

def find_max_key(info_about_T):
    max_path = -1
    key_max = -1
    for i, j in info_about_T.items():
        if j > max_path:
            max_path = j
            key_max = i
    return key_max

def find_max_path(connection_from_user, info_about_T):
   key_max = find_max_key(info_about_T)
   value_to_next = int(key_max.split('_')[1])
   prom_path = [value_to_next]

   while connection_from_user[value_to_next][0] != -1:
        prom = [[i, info_about_T.get(f'T_{i}')] for i in connection_from_user[value_to_next]]
        max_element = -1
        number_info = -1
        for i in prom:
            if i[1] > max_element:
                max_element = i[1]
                number_info = i[0]
        prom_path.append(number_info)
        value_to_next = number_info

   return prom_path
   #      print('-----')
   #      print(connection_from_user[value_to_next])
   #
   #      print('-----')
   #
   # print(prom_path)
from math import ceil, floor
def opt_v1_without_change_path(connection_from_user, ranked_info, info_about_work,info_about_faster_work, max_path, alfa, money, max_user_time):
    dict_values = {}
    # x = cp.Variable()
    print(max_path)
    constraints = []
    string_model = []
    for i in max_path:
        dict_values[f'x_{i}'] = cp.Variable()
        string_model.append(f"{info_about_work[i]}*(1-{alfa[i]}*x_{i}) >= {info_about_faster_work[i]}")
        constraints.append(info_about_work[i]*(1-alfa[i]*dict_values.get(f'x_{i}'))>= info_about_faster_work[i])

    # constraints = [x  < 10]

    # constraints = [dict_values.get(f'x_{7}')]
    # print(constraints)
    # # print(alfa)
    # print([info_about_work[i] for i in max_path])
    # # print([dict_values.get(f'x_{i}') * 1 < 10 for i in max_path])
    #
    prom_text = ''
    for i in max_path:
        prom_text += f'x_{i} + '

    string_model.append(prom_text[:-2] + f'<= {max_user_time}')
    constraints.append(sum([info_about_work[i]*(1-alfa[i]*dict_values.get(f'x_{i}')) for i in max_path]) <= max_user_time)

    string_model.append(prom_text[:-2] + f'<= {money}')
    constraints.append(sum([ dict_values.get(f'x_{i}') for i in max_path]) <= money)

    for i in max_path:
        string_model.append(f'x_{i} >= 0')
        constraints.append(dict_values.get(f'x_{i}') >= 0)
    # print('\ncontains')
    # for i in range(len(constraints)):
    #
    #     print(str(constraints[i]))
    # print('contains\n')

    # for i in max_path:
    #     prom =
    #
    string_model.append(prom_text+" -> min")

    obj = cp.Minimize(sum([dict_values.get(f'x_{i}') for i in max_path]))
    #
    prob = cp.Problem(obj, constraints)
    #
    prob.solve()
    print("status:", prob.status)

    if str(prob.status) == 'optimal':
        print("optimal val:", np.round(prob.value, 5))
        new_t = []
        x_i = []
        for i in max_path:
            print("optimal var {0}: (floor){1}\n".format( f'x_{i}', round(float(dict_values.get(f'x_{i}').value), 3)))
            new_t.append(round(info_about_work[i]*(1-alfa[i]*float(dict_values.get(f'x_{i}').value)), 3))
            x_i.append(round(float(dict_values.get(f'x_{i}').value),3))
            print(round(info_about_work[i] * (1 - alfa[i] * dict_values.get(f'x_{i}').value), 3))
        # print(x_i)
        # print(new_t)
        return [x_i, string_model, new_t, prob.value]
    else:
        return [404, string_model]

def opt_v2_math_full_model(connection_from_user, ranked_info,info_about_work,info_about_faster_work,max_path, alfa, money,max_user_time, info_about_t_and_T):
    dict_values = {}
    # x = cp.Variable()
    # print(ranked_info)
    constraints = []
    prom_text = []
    for i in range(len(connection_from_user)):
        dict_values[f'x_{i}'] = cp.Variable()
        dict_values[f't_{i}'] = cp.Variable()
        prom_text.append(f"x_{i} >= 0")
        constraints.append(dict_values.get(f'x_{i}') >= 0)
    dict_values[f'T_k'] = cp.Variable()
    # for i in
    # print(info_about_work)
    for i, j in ranked_info.items():
        if i == -1:
            for i_1 in j:
                constraints.append(dict_values.get(f't_{i_1}') == 0)

            pass
        elif i == 0:
            for i_1 in j:
                # print(connection_from_user[i_1])
                for i_2 in connection_from_user[i_1]:
                    prom_text.append( f"t{i_1+1}>= {info_about_work[i_2]}*(1-{alfa[i_2]}*x{i_2+1})")
                    # print( f"t{i_1+1}>= {info_about_work[i_2]}*(1-{alfa[i_2]}*x{i_2+1})")
                    constraints.append(
                        dict_values.get(f't_{i_1}') >= info_about_work[i_2]*(1-alfa[i_2]*dict_values.get(f'x_{i_2}')))
            # print()
        else:
            for i_1 in j:
                # print(connection_from_user[i_1])
                for i_2 in connection_from_user[i_1]:
                    prom_text.append(f"t{i_1+1} >= t{i_2+1} + {info_about_work[i_2]}*(1-{alfa[i_2]}*x{i_2+1})")

                    # print( f"t{i_1+1} >= t{i_2+1} + {info_about_work[i_2]}*(1-{alfa[i_2]}*x{i_2+1})")

                    constraints.append(
                        dict_values.get(f't_{i_1}') >= dict_values.get(f't_{i_2}') +  info_about_work[i_2] * (1 - alfa[i_2] * dict_values.get(f'x_{i_2}')))
        # else:
        #     for i_1 in j:
        #         # print(connection_from_user[i_1])
        #         for i_2 in connection_from_user[i_1]:
    # print()
    for i in range(len(connection_from_user)):
        prom_text.append(f"Tk>= {info_about_work[i]}*(1-{alfa[i]}*x{i+1})")
        # print(f"Tk>= {info_about_work[i]}*(1-{alfa[i]}*x{i+1})")

        constraints.append(dict_values.get(f'T_k') >= dict_values.get(f't_{i}') + info_about_work[i] * (1 - alfa[i] * dict_values.get(f'x_{i}')))
    # print()
    for i in range(len(connection_from_user)):
        prom_text.append(f"{info_about_work[i]}*(1-{alfa[i]}*x{i+1}) >= {info_about_faster_work[i]}")
        # print(f"{info_about_work[i]}*(1-{alfa[i]}*x{i+1}) >= {info_about_faster_work[i]}")

        constraints.append(info_about_work[i] * (1 - alfa[i] * dict_values.get(f'x_{i}')) >=  info_about_faster_work[i])

    prom_text.append(f"{dict_values.get(f'T_k')} <= {max_user_time}")
    constraints.append(dict_values.get(f'T_k') <= max_user_time)

    prom_text_val = ''
    for i in range(len(connection_from_user)):
        prom_text_val += f'x_{i} + '
    prom_text_val = prom_text_val[:-2]
    prom_text.append(prom_text_val+ f" <= {money}")
    constraints.append(sum([dict_values.get(f'x_{i}') for i in range(len(connection_from_user))]) <= money)

    prom_text.append(prom_text_val+ f" -> min")

    obj = cp.Minimize(sum([dict_values.get(f'x_{i}') for i in range(len(connection_from_user))]))

    prob = cp.Problem(obj, constraints)
    #
    prob.solve()
    print("status:", prob.status)
    print("optimal val:", np.round(prob.value, 5))
    if str(prob.status) == 'optimal':
        k = 0
        new_t = []
        x_i = []
        new_T = []
        new_tao = []
        for i in range(len(connection_from_user)):
            # print("optimal var {0}: (floor){1}\n".format(f'x_{i}', round(float(dict_values.get(f'x_{i}').value), 3)))
            new_t.append(round(float(info_about_work[i] * (1 - alfa[i] * float(dict_values.get(f'x_{i}').value))), 3))
            x_i.append(round(float(dict_values.get(f'x_{i}').value), 3))
            # print(f't_{i}', round(float(dict_values.get(f't_{i}').value),3))
            new_tao.append(round(float(dict_values.get(f't_{i}').value), 3))
            new_T.append(round(float(dict_values.get(f't_{i}').value) + info_about_work[i] * (1 - alfa[i] * float(dict_values.get(f'x_{i}').value)), 3))
            # print(round(info_about_work[i] * (1 - alfa[i] * dict_values.get(f'x_{i}').value), 2))
        # print(x_i)
        # print(new_t)
        print('-----')
        print(prom_text)
        print('-----')
        return [new_tao, new_t, new_T, x_i, prob.value, prom_text]
    else:
        return [404, prom_text]


def find_connection_for_opt_v3(connection_from_user, ranked_info, info_about_work, info_about_faster_work, max_path, alfa, money, max_user_time,info_about_t_and_T):
    dict_paths = {}
    prom = []
    print('fsdfs')
    for i in reversed(max_path):
        dict_paths[i] = []
        for i_1, j_1 in info_about_t_and_T.items():
            if info_about_t_and_T.get(f'T_{i}') >= j_1 and 'T' in i_1 and f'T_{i}' != i_1 and not int(i_1.split('_')[1]) in max_path  :
                if not i_1 in prom:
                    prom.append(i_1)
                    dict_paths[i].append(i_1)
    print('fsfds')
    for i, j in dict_paths.items():
        print(i,'->',j)

def find_all_paths_T(connection_from_user, ranked_info, info_about_work):
    prom_info_date = []
    dict_info_t = {}
    k = 0

    for i in ranked_info.values():
        prom_info = []
        prom_list_2 = []
        for j in i:
            if connection_from_user[j][0] == -1:

                dict_info_t[f't_{j}'] = [info_about_work[j], [j]]
                dict_info_t[f'T_{j}'] = [info_about_work[j], [j]]
                k += 1
            else:
                # print('---------')
                # print(i)
                # print(j)
                # print(dict_info_t)
                # print(connection_from_user[j])
                # print([i for i in connection_from_user[j]])
                prom_list_2 = []
                for i in connection_from_user[j]:
                    for x in dict_info_t.get(f'T_{i}')[1]:
                        prom_list_2.append(x)

                # print([x for x in dict_info_t.get(f'T_{i}') for i in connection_from_user[j]])
                dict_info_t[f't_{j}'] = [info_about_work[j], [j]]
                prom_list = [dict_info_t.get(f'T_{i}')[0] for i in connection_from_user[j]]
                #
                # print(prom_list_2)
                # print('---------')

                # # for i in connection_from_user:
                # if f'T_{k}' == 'T_11':
                #     print('------')
                #     print(dict_info_mini_t)
                #     print(connection_from_user[j])
                #     print(prom_list)
                #     print('------')
                prom_list_2.append(j)
                dict_info_t[f'T_{j}'] = [info_about_work[j] + max(prom_list), copy.deepcopy(prom_list_2)]
                k += 1
        prom_list_2.clear()
        prom_info_date.append(copy.deepcopy(prom_info))
        prom_info.clear()
    # print(dict_info_big_T)
    # print(dict_info_mini_t)
    for i, j in dict_info_t.items():
        if 'T' in i:
            # print(i,'->', set(j[1]))
            dict_info_t[i] = [j[0], set(j[1])]
    print(dict_info_t)
    return dict_info_t


        # prom = sum([ranked_info[x] for x in j_1 for j_1 in info_about_t_and_T.get(f'T_{i}')])
        # print(prom)

        # for i_1, j_1 in info_about_t_and_T.items():
        #     if

def find_sub_paths(connection_from_user, max_path):
    print(connection_from_user[3])
    for i in reversed(max_path):
        print(i)


def opt_v4_math_full_model(connection_from_user, ranked_info, info_about_work, info_about_faster_work, max_path, alfa, money, max_user_time,info_about_t_and_T):
    dict_values = {}
    # x = cp.Variable()
    # print(ranked_info)
    constraints = []
    string_model = []
    for i in range(len(connection_from_user)):
        dict_values[f'x_{i}'] = cp.Variable()
        dict_values[f't_{i}'] = cp.Variable()
        # constraints.append(dict_values.get(f'x_{i}') >= 0)
    dict_values[f'T_k'] = cp.Variable()

    # print(info_about_work)
    for i, j in ranked_info.items():
        if i == -1:
            pass
        elif i == 0:
            for i_1 in j:
                # print(connection_from_user[i_1])
                for i_2 in connection_from_user[i_1]:
                    string_model.append(f"t{i_1+1}>= {info_about_work[i_2]}*(1-{alfa[i_2]}*x{i_2+1})")
                    # print( f"t{i_1+1}>= {info_about_work[i_2]}*(1-{alfa[i_2]}*x{i_2+1})")
                    constraints.append(
                        dict_values.get(f't_{i_1}') >= info_about_work[i_2]*(1-alfa[i_2]*dict_values.get(f'x_{i_2}')))
            # print()
        else:
            for i_1 in j:
                # print(connection_from_user[i_1])
                for i_2 in connection_from_user[i_1]:
                    string_model.append(f"t{i_1+1} >= t{i_2+1} + {info_about_work[i_2]}*(1-{alfa[i_2]}*x{i_2+1})")

                    # print( f"t{i_1+1} >= t{i_2+1} + {info_about_work[i_2]}*(1-{alfa[i_2]}*x{i_2+1})")

                    constraints.append(
                        dict_values.get(f't_{i_1}') >= dict_values.get(f't_{i_2}') +  info_about_work[i_2] * (1 - alfa[i_2] * dict_values.get(f'x_{i_2}')))
        # else:
        #     for i_1 in j:
        #         # print(connection_from_user[i_1])
        #         for i_2 in connection_from_user[i_1]:
    # print()
    for i in range(len(connection_from_user)):
        string_model.append(f"Tk>= {info_about_work[i]}*(1-{alfa[i]}*x{i+1})")

        # print(f"Tk>= {info_about_work[i]}*(1-{alfa[i]}*x{i+1})")

        constraints.append(dict_values.get(f'T_k') >= dict_values.get(f't_{i}') + info_about_work[i] * (1 - alfa[i] * dict_values.get(f'x_{i}')))
    # print()
    for i in range(len(connection_from_user)):
        string_model.append(f"{info_about_work[i]}*(1-{alfa[i]}*x{i+1}) >= {info_about_faster_work[i]}")

        # print(f"{info_about_work[i]}*(1-{alfa[i]}*x{i+1}) >= {info_about_faster_work[i]}")

        constraints.append(info_about_work[i] * (1 - alfa[i] * dict_values.get(f'x_{i}')) >=  info_about_faster_work[i])

    # constraints.append(dict_values.get(f'T_k') <= max_user_time)
    prom_text_val = ''
    for i in range(len(connection_from_user)):
        prom_text_val += f'x_{i} + '
    prom_text_val = prom_text_val[:-2]
    string_model.append(prom_text_val + ' <= 0 ')
    constraints.append(sum([dict_values.get(f'x_{i}') for i in range(len(connection_from_user))]) <= 0)

    string_model.append('T_k -> min')

    obj = cp.Minimize(dict_values.get(f'T_k'))

    prob = cp.Problem(obj, constraints)
    #
    prob.solve()
    print("status:", prob.status)
    print("optimal val:", np.round(prob.value, 5))
    if str(prob.status) == 'optimal':
        k = 0
        new_t = []
        x_i = []
        new_tao = []
        new_T = []
        for i in range(len(connection_from_user)):
            print("optimal var {0}: (floor){1} ".format(f'x_{i}', round(float(dict_values.get(f'x_{i}').value), 3)))
            new_t.append(round(float(info_about_work[i] * (1 - alfa[i] * float(dict_values.get(f'x_{i}').value))), 3))
            x_i.append(round(float(dict_values.get(f'x_{i}').value), 3))
            # print(f't_{i}', round(float(dict_values.get(f't_{i}').value),3))
            new_tao.append(round(float(dict_values.get(f't_{i}').value), 3))
            new_T.append(round(float(dict_values.get(f't_{i}').value) + info_about_work[i] * (
                        1 - alfa[i] * float(dict_values.get(f'x_{i}').value)), 3))
            # print(round(info_about_work[i] * (1 - alfa[i] * dict_values.get(f'x_{i}').value), 2))
            # print(x_i)
            # print(new_t)
        # print('-----')
        # print(prom_text)
        # print('-----')
        return [new_tao, new_t, new_T, x_i, prob.value, string_model]
    else:
        return [404, string_model]

#
def opt_v7_math_full_model(connection_from_user, ranked_info, info_about_work, info_about_faster_work, max_path, alfa, money, max_user_time,info_about_t_and_T):
    dict_values = {}
    # x = cp.Variable()
    # print(ranked_info)
    constraints = []
    string_model = []
    for i in range(len(connection_from_user)):
        dict_values[f'x_{i}'] = cp.Variable()
        dict_values[f't_{i}'] = cp.Variable()
        string_model.append(f'x_{i} <= 0')
        constraints.append(dict_values.get(f'x_{i}') <= 0)
    dict_values[f'T_k'] = cp.Variable()

    # print(info_about_work)
    for i, j in ranked_info.items():
        if i == -1:
            pass
        elif i == 0:
            for i_1 in j:
                # print(connection_from_user[i_1])
                for i_2 in connection_from_user[i_1]:
                    string_model.append(f"t{i_1+1}>= {info_about_work[i_2]}*(1-{alfa[i_2]}*x{i_2+1})")

                    print( f"t{i_1+1}>= {info_about_work[i_2]}*(1-{alfa[i_2]}*x{i_2+1})")
                    constraints.append(
                        dict_values.get(f't_{i_1}') >= info_about_work[i_2] * (
                                    1 - alfa[i_2] * dict_values.get(f'x_{i_2}')))
            # print()
        else:
            for i_1 in j:
                # print(connection_from_user[i_1])
                for i_2 in connection_from_user[i_1]:
                    string_model.append( f"t{i_1+1} >= t{i_2+1} + {info_about_work[i_2]}*(1-{alfa[i_2]}*x{i_2+1})")

                    print( f"t{i_1+1} >= t{i_2+1} + {info_about_work[i_2]}*(1-{alfa[i_2]}*x{i_2+1}),")

                    constraints.append(
                        dict_values.get(f't_{i_1}') >= dict_values.get(f't_{i_2}') + info_about_work[i_2] * (
                                    1 - alfa[i_2] * dict_values.get(f'x_{i_2}')))
        # else:
        #     for i_1 in j:
        #         # print(connection_from_user[i_1])
        #         for i_2 in connection_from_user[i_1]:
    # print()
    for i in range(len(connection_from_user)):
        string_model.append(f"Tk>= t{i+1} +{info_about_work[i]}*(1-{alfa[i]}*x{i+1})")

        print(f"Tk>= t{i+1} +{info_about_work[i]}*(1-{alfa[i]}*x{i+1}),")

        constraints.append(dict_values.get(f'T_k') >= dict_values.get(f't_{i}') + info_about_work[i] * (
                    1 - alfa[i] * dict_values.get(f'x_{i}')))
    # print()
    # for i in range(len(connection_from_user)):
    #     print(f"{info_about_work[i]}*(1-{alfa[i]}*x{i+1}) >= {info_about_faster_work[i]},")
    #
    #     constraints.append(info_about_work[i] * (1 - alfa[i] * dict_values.get(f'x_{i}')) >= info_about_faster_work[i])
    string_model.append(f"T_k == {max_user_time}")

    constraints.append(dict_values.get(f'T_k') == max_user_time)

    # constraints.append(sum([dict_values.get(f'x_{i}') for i in range(len(connection_from_user))]) <= money)
    prom_text_val = ''
    for i in range(len(connection_from_user)):
        prom_text_val += f'x_{i} + '
    prom_text_val = prom_text_val[:-2]
    string_model.append(prom_text_val+' -> min')
    obj = cp.Minimize(sum([dict_values.get(f'x_{i}') for i in range(len(connection_from_user))]))

    prob = cp.Problem(obj, constraints)
    #
    prob.solve()
    print("status:", prob.status)
    print("optimal val:", np.round(prob.value, 5))
    if str(prob.status) == 'optimal':
        k = 0
        new_t = []
        x_i = []
        new_tao = []
        new_T = []
        for i in range(len(connection_from_user)):
            print("optimal var {0}: (floor){1} ".format(f'x_{i}', round(float(dict_values.get(f'x_{i}').value), 3)))
            new_t.append(round(float(info_about_work[i] * (1 - alfa[i] * float(dict_values.get(f'x_{i}').value))), 3))
            x_i.append(round(float(dict_values.get(f'x_{i}').value), 3))
            # print(f't_{i}', round(float(dict_values.get(f't_{i}').value),3))
            new_tao.append(round(float(dict_values.get(f't_{i}').value), 3))
            new_T.append(round(float(dict_values.get(f't_{i}').value) + info_about_work[i] * (
                    1 - alfa[i] * float(dict_values.get(f'x_{i}').value)), 3))
            # print(round(info_about_work[i] * (1 - alfa[i] * dict_values.get(f'x_{i}').value), 2))
            # print(x_i)
            # print(new_t)
        # print('-----')
        # print(prom_text)
        # print('-----')
        return [new_tao, new_t, new_T, x_i, prob.value, string_model]
    else:
        return [404, string_model]


def main(connection_from_user, info_about_work, info_about_faster_work, alfa, money, max_user_time, min_user_time, file_name):
    ranked_info = determite_work(connection_from_user)
    print('info_about_t_and_T')
    print('coast' , info_about_work)
    info_about_t_and_T = find_t_and_T(connection_from_user, ranked_info, info_about_work)
    print('find_max_path')
    max_path = find_max_path(connection_from_user, info_about_t_and_T)
    # print('opt1')
    # find_sub_paths(connection_from_user, max_path)
    opt_1 = opt_v1_without_change_path(connection_from_user, ranked_info, info_about_work, info_about_faster_work, max_path, alfa, money, max_user_time)
    # print('opt2')
    opt_2 = opt_v2_math_full_model(connection_from_user, ranked_info, info_about_work, info_about_faster_work, max_path, alfa, money, max_user_time, info_about_t_and_T)
    # print('opt4')

    opt_4 = opt_v4_math_full_model(connection_from_user, ranked_info, info_about_work, info_about_faster_work, max_path, alfa, money, max_user_time, info_about_t_and_T)
    # print('opt7')
    opt_7 = opt_v7_math_full_model(connection_from_user, ranked_info, info_about_work, info_about_faster_work, max_path, alfa, money, min_user_time, info_about_t_and_T)
    to_excel_data(file_name,connection_from_user, ranked_info, info_about_work, info_about_faster_work, alfa, opt_1, opt_2, opt_4, opt_7,money, max_user_time, min_user_time, max_path)

def to_excel_data(file_name,connection_from_user,ranked_info,  info_about_work, info_about_faster_work, alfa, opt_1, opt_2, opt_4, opt_7,money, max_user_time, min_user_time,max_path):
    # print(connection_from_user)
    # print(ranked_info)
    wb2 = load_workbook('Книга1.xlsx', data_only=True)
    ws4 = wb2["Ранги"]
    # ws4['K3'].value = 'Успешно' if opt_1[0] != 404 else 'Неуспешно'
    ws4['K3'].value = 'Успешно' if opt_2[0] != 404 else 'Неуспешно'
    ws4['K4'].value = 'Успешно' if opt_4[0] != 404 else 'Неуспешно'
    ws4['K5'].value = 'Успешно' if opt_7[0] != 404 else 'Неуспешно'

    # ws4['K3'].fill = PatternFill(fill_type='solid', start_color='2A6FD7') if opt_1[0] != 404 else  PatternFill(fill_type='solid', start_color='FF0000')
    ws4['K3'].fill = PatternFill(fill_type='solid', start_color='2A6FD7') if opt_2[0] != 404 else  PatternFill(fill_type='solid', start_color='FF0000')
    ws4['K4'].fill = PatternFill(fill_type='solid', start_color='2A6FD7') if opt_4[0] != 404 else  PatternFill(fill_type='solid', start_color='FF0000')
    ws4['K5'].fill = PatternFill(fill_type='solid', start_color='2A6FD7') if opt_7[0] != 404 else  PatternFill(fill_type='solid', start_color='FF0000')

    ws4['K3'].font = Font(size=12, color='FFFFFF', bold=True, italic=True)
    ws4['K4'].font = Font(size=12, color='FFFFFF', bold=True, italic=True)
    ws4['K5'].font = Font(size=12, color='FFFFFF', bold=True, italic=True)
    # ws4['K6'].font = Font(size=12, color='FFFFFF', bold=True, italic=True)

    ws4['K3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws4['K4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws4['K5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # ws4['K6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)




    # ws4['J4'].font = Font(size=11, color='00595959', b=False)


    ranked_excel(ws4, ranked_info, connection_from_user, info_about_work, info_about_faster_work, alfa, money, max_user_time, min_user_time,max_path)
    # ws5 = wb2['ОПТ №1']
    # opt_v1_excel(ws5, opt_1)

    ws6 = wb2['ОПТ №1']
    opt_v2_excel(ws6, opt_2)

    ws7 = wb2['ОПТ №2']
    opt_v4_excel(ws7, opt_4)

    ws8 = wb2['ОПТ №3']
    opt_v7_excel(ws8, opt_7)

    wb2.save(f'{file_name}.xlsx')

def ranked_excel(ws4,ranked_info,connection_from_user, info_about_work,info_about_faster_work, alfa, money, max_user_time, min_user_time,max_path):
    k = 3
    for i in ranked_info:
        for j in ranked_info.get(i):
            print(j)
            ws4[f'A{k}'].value = k - 2
            ws4[f'B{k}'].value = j
            if i == -1:
                ws4[f'C{k}'].value = '-'
            else:
                ws4[f'C{k}'].value = ', '.join([str(i) for i in connection_from_user[j]])
            ws4[f'D{k}'].value = i + 1
            ws4[f'E{k}'].value = info_about_work[j]
            ws4[f'F{k}'].value = info_about_faster_work[j]
            ws4[f'G{k}'].value = alfa[j]
            for fot in ['A','B','C','D','E','F','G']:
                ws4[f'{fot}{k}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws4[f'{fot}{k}'].font = Font(size=12)

            k += 1

    k += 1
    ws4[f'B{k}'].value = 'Средства'
    ws4[f'D{k}'].value = money

    ws4[f'B{k}'].font = Font(size=14)
    ws4[f'D{k}'].font = Font(size=14)

    k += 1
    ws4[f'B{k}'].value = 'T max проекта'
    ws4[f'D{k}'].value = min_user_time

    ws4[f'B{k}'].font = Font(size=14)
    ws4[f'D{k}'].font = Font(size=14)

    k += 1
    ws4[f'B{k}'].value = 'T min проекта'
    ws4[f'D{k}'].value = max_user_time

    ws4[f'B{k}'].font = Font(size=14)
    ws4[f'D{k}'].font = Font(size=14)

    ws4[f'O2'].value = sum([info_about_work[i] for i in max_path])


    k = 3
    for i in reversed(max_path):
        ws4[f'M{k}'].value = i
        ws4[f'M{k}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws4[f'M{k}'].font = Font(size=14)

        k += 1


def opt_v1_excel(ws5, opt_1):
    k = 2
    for i in opt_1[1]:
        ws5[f'A{k}'].value = i
        k += 1
    ws5['P1'].fill = PatternFill(fill_type='solid', start_color='2A6FD7') if opt_1[0] != 404 else PatternFill(fill_type='solid', start_color='FF0000')

    if opt_1[0] != 404:
        ws5['P1'].value = 'Успешно'
        ws5['X1'].value = opt_1[3]

        ws5['P1'].font = Font(size=12, color='FFFFFF')


        k = 3
        # print(opt_1[0])
        # print(opt_1[2])

        for i in range(len(opt_1[0])):
            ws5[f'O{k}'].value = opt_1[0][i]
            ws5[f'P{k}'].value = opt_1[2][i]
            k+=1


    else:
        ws5['P1'].value = 'Неуспешно'
        ws5['P1'].font = Font(size=12, color='FFFFFF')


def opt_v2_excel(ws6, opt_2):
    k = 2
    print(opt_2)
    for i in opt_2[-1]:
        ws6[f'A{k}'].value = i
        k += 1
    ws6['P1'].fill = PatternFill(fill_type='solid', start_color='2A6FD7') if opt_2[0] != 404 else PatternFill(fill_type='solid', start_color='FF0000')

    if opt_2[0] != 404:
        ws6['P1'].value = 'Успешно'
        ws6['W1'] = opt_2[-2]
        k = 3
        ws6['P1'].font = Font(size=12, color='FFFFFF')
        # print(opt_1[0])
        # print(opt_1[2])

        for i in range(len(opt_2[0])):
            ws6[f'O{k}'].value = k-2
            ws6[f'P{k}'].value = opt_2[0][i]
            ws6[f'Q{k}'].value = opt_2[1][i]
            ws6[f'R{k}'].value = opt_2[2][i]
            ws6[f'S{k}'].value = opt_2[3][i]
            # print(opt_2[4])
            # ws6[f'S{k}'].value = opt_2[4][i]
            k += 1


    else:
        ws6['P1'].value = 'Неуспешно'
        ws6['P1'].font = Font(size=12, color='FFFFFF')

def opt_v4_excel(ws7, opt_4):
    k = 2
    print(opt_4)
    for i in opt_4[-1]:
        ws7[f'A{k}'].value = i
        k += 1
    ws7['P1'].fill = PatternFill(fill_type='solid', start_color='2A6FD7') if opt_4[0] != 404 else PatternFill(fill_type='solid', start_color='FF0000')

    if opt_4[0] != 404:
        ws7['P1'].value = 'Успешно'
        ws7['U1'] = opt_4[-2]
        k = 3
        ws7['P1'].font = Font(size=12, color='FFFFFF')

        # print(opt_1[0])
        # print(opt_1[2])

        for i in range(len(opt_4[0])):
            ws7[f'O{k}'].value = k - 2
            ws7[f'P{k}'].value = opt_4[0][i]
            ws7[f'Q{k}'].value = opt_4[1][i]
            ws7[f'R{k}'].value = opt_4[2][i]
            ws7[f'S{k}'].value = opt_4[3][i]
            # print(opt_2[4])
            # ws6[f'S{k}'].value = opt_2[4][i]
            k += 1
    else:
        ws7['P1'].value = 'Неуспешно'
        ws7['P1'].font = Font(size=12, color='FFFFFF')


def opt_v7_excel(ws8, opt7):
    k = 2
    for i in opt7[-1]:
        ws8[f'A{k}'].value = i
        k += 1
    ws8['P1'].fill = PatternFill(fill_type='solid', start_color='2A6FD7') if opt7[0] != 404 else PatternFill(fill_type='solid', start_color='FF0000')

    if opt7[0] != 404:
        ws8['P1'].value = 'Успешно'
        ws8['X1'] = opt7[-2]
        ws8['P1'].font = Font(size=12, color='FFFFFF')

        k = 3
        # print(opt_1[0])
        # print(opt_1[2])

        for i in range(len(opt7[0])):
            ws8[f'O{k}'].value = k - 2
            ws8[f'P{k}'].value = opt7[0][i]
            ws8[f'Q{k}'].value = opt7[1][i]
            ws8[f'R{k}'].value = opt7[2][i]
            ws8[f'S{k}'].value = opt7[3][i]
            # print(opt_2[4])
            # ws6[f'S{k}'].value = opt_2[4][i]
            k += 1
    else:
        ws8['P1'].value = 'Неуспешно'
        ws8['P1'].font = Font(size=12, color='FFFFFF')



# def plot_graph(list_from_user = -1):
#     list_from_user = [[-1], [-1], [-1], [0, 1], [0, 1, 2], [0, 1, 2], [5], [3, 4, 6]]
#     import networkx as nx
#     import matplotlib.pyplot as plt
#     G = nx.DiGraph()
#     # G.add_edges_from(
#     #     [('A', 'B'), ('A', 'C'), ('D', 'B'), ('E', 'C'), ('E', 'F'),
#     #      ('B', 'H'), ('B', 'G'), ('B', 'F'), ('C', 'G')])
#     prom = []
#     for i in range(len(list_from_user)):
#         if len(list_from_user[i])==1 and list_from_user[i][0] == -1:
#             print('fsdfsfd')
#             pass
#         else:
#             for j in list_from_user[i]:
#                 prom.append(str(i)+str(j))
#     print(prom)
#     dot = Digraph(comment='The Round Table')
#     # dot.node('A', 'King Arthur')
#     # dot.node('B', 'Sir Bedevere the Wise')
#     # dot.node('L', 'Sir Lancelot the Brave')
#
#     dot.edges(reversed(prom))
#     # dot.edge('B', 'L', constraint='false')
#     dot.render('test-output/round-table.gv', view=True)

# plot_graph()
# пример из начала - то, что он дает на первом методе - ок (не было еще параметров)
# main([[-1], [-1], [-1], [0, 1], [1, 2], [1, 3], [2, 4], [3, 4], [6], [5, 7], [8, 9], [9]],
#      [10, 5, 15, 10, 30, 5, 15, 25, 15, 30, 35, 10],
#      [15, 5, 5, 10, 5, 10, 3, 5],
#           [0.1, 0.15, 0.1, 0.2, 0.2, 0.01, 0.05, 0.1],
#           60,
#           32
#      )
#да.
# main([[-1],[-1],[-1],[0],[0],[2],[1, 4, 5], [1, 4, 5],[3, 6], [2], [1, 3, 5,9]],
#      [8, 10, 6, 9, 5, 2, 4, 13, 8, 17, 10])
# main([[-1],[-1],[-1],[0,1],[0,1,2],[0,1,2],[5],[3,4,6]],
#      [20, 10, 5, 30, 10, 15, 10, 10],
#      [15, 5, 5, 10, 5, 10, 3, 3],
#      [0.1, 0.15, 0.1, 0.2, 0.2, 0.01, 0.05, 0.1],
#      10,
#      35)
# main([[-1],[-1],[-1],[0,1],[0,1,2],[0,1,2],[5],[3,4,6]],
#      [20, 10, 5, 30, 10, 15, 10, 10],
#      [15, 5, 5, 10, 5, 10, 3, 3],
#      [0.1, 0.15, 0.1, 0.2, 0.2, 0.01, 0.05, 0.1],
#      80,
#      32)

#опт_v7
main([[-1], [-1], [-1], [0, 1], [0, 1, 2], [0, 1, 2], [5], [3, 4, 6]],
     [20, 10, 5, 30, 10, 15, 10, 10],
     [15, 5, 5, 10, 5, 10, 3, 3],
     [0.1, 0.15, 0.1, 0.2, 0.2, 0.01, 0.05, 0.1],
     60,
     32,
     80,
     'new_doc')
main([[-1],[-1],[-1],[0],[0],[2],[1,4,5],[1,4,5],[3,6],[3],[1,4,5,9]],
     [8,10,6,9,5,2,4,13,8,17,10],
     [3,4,1,1,1,1,1,4,1,6,2],
     [0.6,0.8,0.4,0.6,0.3,0.2,0.3,0.9,0.5,0.10,0.7],
     10,
     25,
     100,
     'max_none')
a = [8,10,6,9,5,2,4,13,8,17,10]
print(a[0]+a[3]+a[9]+a[10])